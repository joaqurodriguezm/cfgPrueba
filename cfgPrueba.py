from openpyxl import Workbook
from openpyxl.styles import Protection
from openpyxl.worksheet.datavalidation import DataValidation
import xlwings as xw

def inicializar_hoja(ws):
    # Escribir Alumno1, Alumno2, ..., Alumno9 en las celdas A2 hasta A11 y protegerlas
    for i in range(2, 12):
        cell = ws[f'A{i}']
        cell.value = f'Alumno{i-1}'
        cell.protection = Protection(locked=True)  # Proteger celdas A2:A11

    # Escribir Pregunta1, Pregunta2, ..., Pregunta6 en las celdas B1 hasta F1 y protegerlas
    for j in range(2, 8):
        cell = ws.cell(row=1, column=j)
        cell.value = f'Pregunta{j-1}'
        cell.protection = Protection(locked=True)  # Proteger celdas B1:F1

    # Crear lógica para preguntas que acepten enteros y decimales, incluyendo negativos (P1)
    dv1 = DataValidation(type="decimal",
                        operator="between",
                        formula1=None,
                        formula2=None,
                        showErrorMessage=True,
                        errorTitle="Entrada inválida",
                        error="Solo se permiten números enteros o decimales con separador ','")

    for row in ws['B2:B11']:  # Solo para Pregunta 1 (Columna B)
        for cell in row:
            dv1.add(cell)

    # Crear lista desplegable para preguntas abiertas simples (P2)
    dv2 = DataValidation(type="list",
                        formula1='"0,1,2"',
                        showErrorMessage=True,
                        errorTitle="Valor Inválido",
                        error="El valor debe ser uno de los seleccionados en la lista.")

    for row in ws['C2:C11']:  # Solo para Pregunta 2 (Columna C)
        for cell in row:
            dv2.add(cell)

    # Crear lógica para preguntas que acepten fracciones (P3)
    dv3 = DataValidation(type="custom",
                        formula1='=AND(ISNUMBER(VALUE(LEFT(D2,FIND("/",D2)-1))),ISNUMBER(VALUE(MID(D2,FIND("/",D2)+1,LEN(D2)-FIND("/",D2)))),COUNTIF(D2,"*/?*")=1)',
                        showErrorMessage=True,
                        error="Solo se permiten fracciones en formato numerador/denominador, ej: 1/2",
                        errorTitle="Entrada inválida")

    for row in ws['D2:D11']:  # Solo para Pregunta 3 (Columna D)
        for cell in row:
            dv3.add(cell)
            cell.number_format = '@'  # Formato de texto

    # Crear lista desplegable para preguntas de selección única (P4)
    dv4 = DataValidation(type="list",
                        formula1='"A, B, C, D, N, N/C"',
                        showErrorMessage=True,
                        errorTitle="Valor Inválido",
                        error="El valor debe ser uno de los seleccionados en la lista.")

    for row in ws['E2:E11']:  # Solo para Pregunta 4 (Columna E)
        for cell in row:
            dv4.add(cell)

    # Crear lógica para preguntas que acepten pares ordenados (P6)
    dv6 = DataValidation(type="custom",
                        formula1='=AND(ISNUMBER(VALUE(LEFT(G2,FIND(";",G2)-1))),ISNUMBER(VALUE(MID(G2,FIND(";",G2)+1,LEN(G2)-FIND(";",G2)))),COUNTIF(G2,"*;?*")=1)',
                        showErrorMessage=True,
                        error="Solo se permiten valores en formato de par ordenado, ej: X;Y",
                        errorTitle="Entrada inválida")
    
    for row in ws['G2:G11']:
        for cell in row:
            dv6.add(cell)
            cell.number_format = '@'  # Formato de texto

    # Agregar validaciones a la hoja
    ws.add_data_validation(dv1)
    ws.add_data_validation(dv2)
    ws.add_data_validation(dv3)
    ws.add_data_validation(dv4)
    ws.add_data_validation(dv6)

    # Desbloquear solo las celdas B2:G11 (Preguntas 1 a 6)
    for row in ws_preguntas['B2:G11']:
        for cell in row:
            cell.protection = Protection(locked=False)

    # Bloquear las demás celdas por defecto
    ws.protection.sheet = True 

# Crear un archivo Excel
wb = Workbook()

# Inicializar la hoja principal
ws_preguntas = wb.active
ws_preguntas.title = 'Preguntas'
inicializar_hoja(ws_preguntas)

ws_preguntas.column_dimensions['F'].width = 36  # Ajustar ancho de columna F (Pregunta 5)

# Crear e inicializar una nueva hoja
ws_respuestas = wb.create_sheet(title='Respuestas')
inicializar_hoja(ws_respuestas)
ws_respuestas.protection.sheet = True

# Referenciar los datos de B2:E11 en la hoja 'Respuestas', pero dejando vacía si la celda original está vacía
for i in range(2, 12):
    for j in range(2, 6):
        cell_respuesta = ws_respuestas.cell(row=i, column=j)
        cell_pregunta = ws_preguntas.cell(row=i, column=j)
        cell_respuesta.value = f'=IF(Preguntas!{cell_pregunta.coordinate}="","",Preguntas!{cell_pregunta.coordinate})'

# Referenciar los datos de la Pregunta 6 (columna G) en la hoja 'Respuestas'
for i in range(2, 12):  # Para las filas 2 a 11
    cell_respuesta = ws_respuestas.cell(row=i, column=7)  # Columna G en la hoja de respuestas
    cell_pregunta = ws_preguntas.cell(row=i, column=7)    # Columna G en la hoja de preguntas
    cell_respuesta.value = f'=IF(Preguntas!{cell_pregunta.coordinate}="","", "("&Preguntas!{cell_pregunta.coordinate}&")")'

# Referenciar los datos de K1:K10 en la hoja 'Datos' a F2:F11 en la hoja 'Respuestas'
for i in range(2, 12):
    cell_respuesta = ws_respuestas.cell(row=i, column=6)  
    cell_dato = ws_preguntas.cell(row=i-1, column=11)  
    cell_respuesta.value = f"=Datos!K{i-1}"

# Crear la hoja de enlaces para CheckBoxes
ws_datos = wb.create_sheet(title='Datos')

# Guardar el archivo Excel
file_path = r'C:\Users\joaquin.rodriguezm\Desktop\cfgPrueba.xlsx'
wb.save(file_path)

# Usar xlwings para agregar el código VBA que genera los CheckBoxes
app = xw.App(visible=False)  # Crea una instancia de Excel sin mostrar la ventana
wb = app.books.open(file_path)

# Código VBA para agregar las casillas de verificación en el rango F2:F11 y vincularlas a celdas en la hoja Datos
vba_code = """
Sub AddCheckBoxes()
    Dim ws As Worksheet
    Dim wsLinks As Worksheet
    Set ws = ThisWorkbook.Sheets("Preguntas")
    Set wsLinks = ThisWorkbook.Sheets("Datos")
    
    Dim i As Integer
    Dim j As Integer
    Dim leftPos As Double
    Dim topPos As Double
    Dim checkBoxWidth As Double
    Dim labels As Variant
    checkBoxWidth = 40
    
    ' Etiquetas para los checkboxes
    labels = Array("OP1", "OP2", "OP3", "OP4", "OP5")
    
    ' Crear los checkboxes en cada celda en el rango F2:F11
    For i = 2 To 11
        topPos = ws.Cells(i, 6).Top
        leftPos = ws.Cells(i, 6).Left
        
        ' Crear CheckBoxes para cada celda en el rango
        For j = LBound(labels) To UBound(labels)
            With ws.CheckBoxes.Add(leftPos + j * checkBoxWidth, topPos, 20, 15) ' Espaciado horizontal
                .Caption = labels(j)  ' Texto al lado del checkbox
                .Value = xlOff
                
                ' Establecer la celda de enlace en la hoja CheckBoxLinks
                .LinkedCell = wsLinks.Cells(i - 1, j + 1).Address(External:=True)
            End With
        Next j
    Next i
    
    ' Convertir los valores VERDADERO/FALSO en 1/0 en las celdas F1:J10
    For i = 1 To 10
        For j = 1 To 5
            wsLinks.Cells(i, j + 5).Formula = "=IF(" & wsLinks.Cells(i, j).Address & ",1,0)"
        Next j
        
        ' Concatenar los valores en la columna K
        wsLinks.Cells(i, 11).Formula = "=" & wsLinks.Cells(i, 6).Address & "&" & wsLinks.Cells(i, 7).Address & "&" & wsLinks.Cells(i, 8).Address & "&" & wsLinks.Cells(i, 9).Address & "&" & wsLinks.Cells(i, 10).Address
    Next i
End Sub
"""

# Agregar el código VBA al módulo
vba_module = wb.api.VBProject.VBComponents.Add(1)  # 1 representa un módulo estándar
vba_module.CodeModule.AddFromString(vba_code)

# Ejecutar la macro para agregar los checkboxes y establecer los vínculos
wb.api.Application.Run("AddCheckBoxes")

# Guardar y cerrar el archivo
wb.save()
wb.close()
app.quit()
